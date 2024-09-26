#!/usr/bin/env python3

"""
This retrieval script is intended to fetch, clean, and translate wahapedia data
into raw json data. From there it can be ingested into a system of choice.
"""

import argparse
import logging
from pathlib import Path
from typing import Union, List, Dict
import requests
import openpyxl
import csv
import json

# Constants
WAHAPEDIA_URL = "https://wahapedia.ru/wh40k10ed/Export%20Data%20Specs.xlsx"
INDEX_FILENAME = "Index.xlsx"
CSV_EXTENSION = ".csv"
JSON_EXTENSION = ".json"
CLEANED_EXTENSION = ".csv.cleaned"
CSV_DELIMITER = '|'
ENCODING = 'utf-8-sig'

logging.basicConfig(format="%(asctime)s %(levelname)-8s %(message)s")
logger = logging.getLogger(__name__)

def configure_logging(loglevel: Union[int, str]) -> None:
    """Configure logging level."""
    logging.basicConfig(level=loglevel)
    logging.getLogger().setLevel(loglevel)

def req_to_file(req_str: str, directory: Union[str, Path], filename: str) -> None:
    """
    Download content from a URL and save it to a file.

    Args:
        req_str (str): The URL to download from.
        directory (Union[str, Path]): The directory to save the file in.
        filename (str): The name of the file to save.

    Raises:
        requests.RequestException: If the download fails.
        IOError: If writing to the file fails.
    """
    try:
        with requests.get(req_str, stream=True) as req:
            req.raise_for_status()
            file_path = Path(directory) / filename
            with open(file_path, 'wb') as ofile:
                for chunk in req.iter_content(chunk_size=8192):
                    ofile.write(chunk)
        logger.info(f"Successfully downloaded {filename}")
    except requests.RequestException as e:
        logger.error(f"Failed to download {req_str}: {e}")
        raise
    except IOError as e:
        logger.error(f"Failed to write to {file_path}: {e}")
        raise

def retrieve_data(prefix: str, dstdir: Union[str, Path]) -> None:
    """
    Retrieve data from Wahapedia and save to destination directory.

    Args:
        prefix (str): Prefix for saved files.
        dstdir (Union[str, Path]): Destination directory for saved files.
    """
    dstdir = Path(dstdir)
    req_to_file(WAHAPEDIA_URL, dstdir, INDEX_FILENAME)

    try:
        wb = openpyxl.load_workbook(dstdir / INDEX_FILENAME)
        ws = wb['EN']

        for row in ws.iter_rows(min_col=1, min_row=ws.min_row, max_col=1, max_row=ws.max_row):
            for cell in row:
                if cell.value and cell.value.endswith(CSV_EXTENSION):
                    logger.info(f"Retrieving {cell.value} from {cell.hyperlink.target}")
                    filename = f"{prefix}-{cell.value}" if prefix else cell.value
                    req_to_file(cell.hyperlink.target, dstdir, filename)
    except Exception as e:
        logger.error(f"Error processing Excel file: {e}")
        raise

def convert_csv_to_json(srcdir: Union[str, Path], dstdir: Union[str, Path]) -> None:
    """
    Convert CSV files to JSON format.

    Args:
        srcdir (Union[str, Path]): Source directory containing CSV files.
        dstdir (Union[str, Path]): Destination directory for JSON files.
    """
    srcdir, dstdir = Path(srcdir), Path(dstdir)
    for csv_file in srcdir.glob(f'*{CSV_EXTENSION}'):
        logger.info(f"Processing file: {csv_file}...")

        try:
            with open(csv_file, 'r', encoding=ENCODING) as csvf:
                csvReader = csv.DictReader(csvf, delimiter=CSV_DELIMITER)
                jsonArray: List[Dict[str, str]] = []

                for row in csvReader:
                    if '' in row:
                        del row['']
                    jsonArray.append(row)

            json_file = dstdir / f"{csv_file.stem}{JSON_EXTENSION}"
            with open(json_file, 'w', encoding='utf-8') as jsonf:
                logger.info(f"Writing post-processed file: {json_file}...")
                json.dump(jsonArray, jsonf, indent=4)

        except Exception as e:
            logger.error(f"Error processing {csv_file}: {e}")

def main() -> None:
    """Main function to handle command line arguments and execute tasks."""
    parser = argparse.ArgumentParser(description='Fetches csv data files from wahapedia')
    parser.add_argument("-c", "--convert", action='store_true',
                        help="Attempts to convert all files in dstdir from wahapedia csv (pipe delimited format) into json format")
    parser.add_argument("-f", "--fetch", action='store_true',
                        help="Attempts to fetch all files from wahapedia index file")
    parser.add_argument("-l", "--loglevel", type=str,
                        help="set log level")
    parser.add_argument("-o", "--out_dir", type=str,
                        help="directory into which to place the files", default=".")
    parser.add_argument("-s", "--src_dir", type=str,
                        help="directory from which to take files for conversion", default=".")

    args = parser.parse_args()

    if args.loglevel:
        caps_log_level = args.loglevel.upper()
        numeric_level = getattr(logging, caps_log_level, None)
        if not isinstance(numeric_level, int):
            raise ValueError(f'Invalid log level: {args.loglevel}')
        configure_logging(numeric_level)
        logger.info(f"Logging Level:{caps_log_level} numeric:{numeric_level}")

    if args.fetch:
        retrieve_data("", args.out_dir)
        args.src_dir = args.out_dir

    if args.convert:
        convert_csv_to_json(args.src_dir, args.out_dir)

if __name__ == "__main__":
    main()
    # RUN: python get_datasheets.py -f -c -o ../wahapedia_data -s ../wahapedia_data